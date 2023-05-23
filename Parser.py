import xml.etree.ElementTree as ET
from openpyxl import Workbook

# Parse the XML file
tree = ET.parse('Sourcefiles/MF3.xml')

# Get the root element
root = tree.getroot()

image_urls = []

# Iterate over each 'item' element in the XML file
for item in root.findall('./channel/item'):
    # Get the 'content' element
    url = item.find('guid').text
    urlsplit=url.rsplit('/', 1)[-1]



    # If the 'content' element exists and has a 'url' attribute
    if url is not None:
        # Add the URL to the list of image URLs
        image_urls.append(urlsplit)

# Create a new Excel workbook and sheet
wb = Workbook()
ws = wb.active

# Write the image URLs to the Excel sheet
for i, urlsplit  in enumerate(image_urls):
    ws.cell(row=i+1, column=1, value=urlsplit)
# Define a list to store the image URLs
image_urls = []

# Iterate over each 'item' element in the XML file
for item in root.findall('./channel/item'):
    # Get the 'content' element
    url = item.find('guid').text

    # If the 'content' element exists and has a 'url' attribute
    if url is not None:
        # Add the URL to the list of image URLs
        image_urls.append(url)


# Create a new Excel workbook and sheet

# Write the image URLs to the Excel sheet
for i, url in enumerate(image_urls):
    ws.cell(row=i+1, column=2, value=url)

# Save the Excel file
wb.save('Sourcefiles/Bildurl.xlsx')

