import openpyxl

def verketten():
    # Öffnen der Eingabe-XLSX-Datei und Laden des Arbeitsblatts
    workbook = openpyxl.load_workbook('Source/Kurzbeschreibung.xlsx')

    worksheet = workbook.active

    # Kopieren der Werte von Spalte C in Spalte A
    prev_value = None  # Speichern des vorherigen Werts in Spalte B
    col_a_value = None  # Initialisieren des Werts von col_a_value
    current_a_row = 2  # Starten der Schleife ab der 2. Zeile
    current_c_row = 3  # Starten der Schleife ab der 2. Zeile
    last_row = worksheet.max_row
    worksheet.cell(row=2, column=2).value = str(worksheet.cell(row=2, column=4).value)
    while current_c_row <= last_row:
        col_a_value = worksheet.cell(row=current_a_row, column=1).value
        col_c_value = worksheet.cell(row=current_c_row, column=3).value
        col_b_value = worksheet.cell(row=current_a_row, column=2).value
        col_d_value = worksheet.cell(row=current_c_row, column=4).value
        if col_c_value is col_a_value:
            worksheet.cell(row=current_a_row, column=2).value = str(worksheet.cell(row=current_a_row, column=2).value) + ", " + str(worksheet.cell(
                row=current_c_row, column=4).value)
            current_c_row += 1
        else:
            current_a_row += 1
            worksheet.cell(row=current_a_row, column=2).value = str(worksheet.cell(row=current_a_row, column=2).value) + ", " + str(col_d_value)
            current_c_row += 1

# Speichern der Änderungen in der XLSX-Datei
    workbook.save('output.xlsx')
    return True
