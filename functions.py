import openpyxl
import xml.etree.ElementTree as ET


def chained():
    workbook = openpyxl.load_workbook('Source/Test_Matrix.xlsx')

    worksheet = workbook.active
    counter=2
    last_row = worksheet.max_row
    while counter <= last_row:
        aktiv = worksheet.cell(row=counter, column=4).value
        set = worksheet.cell(row=counter, column=2).value
        anzahl = 0
        hauptzahl = 2
        while hauptzahl <= last_row:
            if set is worksheet.cell(row=hauptzahl, column=2).value:
                if aktiv is worksheet.cell(row=hauptzahl, column=4).value:
                    anzahl +=1
            hauptzahl += 1
        worksheet.cell(row=counter, column=6).value = anzahl
        counter += 1
    workbook.save('chained.xlsx')
    return True

def verketten():
    # Öffnen der Eingabe-XLSX-Datei und Laden des Arbeitsblatts
    workbook = openpyxl.load_workbook('Source/Kurzbeschreibung.xlsx')

    worksheet = workbook.active

    # Kopieren der Werte von Spalte C in Spalte A
    prev_value = None  # Speichern des vorherigen Werts in Spalte B
    col_a_value = None  # Initialisieren des Werts von col_a_value
    current_a_row = 2  # Starten der Schleife ab der 2. Zeile
    current_c_row = 3  # Starten der Schleife ab der 2. Zeile
    last_row = worksheet.max_row
    worksheet.cell(row=2, column=2).value = str(worksheet.cell(row=2, column=4).value)
    while current_c_row <= last_row:
        col_a_value = worksheet.cell(row=current_a_row, column=1).value
        col_c_value = worksheet.cell(row=current_c_row, column=3).value
        col_b_value = worksheet.cell(row=current_a_row, column=2).value
        col_d_value = worksheet.cell(row=current_c_row, column=4).value
        if col_c_value is col_a_value:
            worksheet.cell(row=current_a_row, column=2).value = str(worksheet.cell(row=current_a_row, column=2).value) + ", " + str(worksheet.cell(
                row=current_c_row, column=4).value)
            current_c_row += 1
        else:
            current_a_row += 1
            worksheet.cell(row=current_a_row, column=2).value = str(worksheet.cell(row=current_a_row, column=2).value) + ", " + str(col_d_value)
            current_c_row += 1

# Speichern der Änderungen in der XLSX-Datei
    workbook.save('output.xlsx')
    return True

def xml_parser():
    # Parse the XML file
    tree = ET.parse('Source/source.xml')

    # Get the root element
    root = tree.getroot()

    image_urls = []

        # Iterate over each 'item' element in the XML file
    for item in root.findall('./channel/item'):
        # Get the 'content' element
        url = item.find('guid').text
        urlsplit=url.rsplit('/', 1)[-1]



        # If the 'content' element exists and has a 'url' attribute
        if url is not None:
            # Add the URL to the list of image URLs
            image_urls.append(urlsplit)

    # Create a new Excel workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active

    # Write the image URLs to the Excel sheet
    for i, urlsplit in enumerate(image_urls):
        ws.cell(row=i+1, column=1, value=urlsplit)
    # Define a list to store the image URLs
    image_urls = []

    # Iterate over each 'item' element in the XML file
    for item in root.findall('./channel/item'):
        # Get the 'content' element
        url = item.find('guid').text

        # If the 'content' element exists and has a 'url' attribute
        if url is not None:
            # Add the URL to the list of image URLs
            image_urls.append(url)


    # Create a new Excel workbook and sheet

    # Write the image URLs to the Excel sheet
    for i, url in enumerate(image_urls):
        ws.cell(row=i+1, column=2, value=url)

    # Save the Excel file
    wb.save('result.xlsx')
    return True