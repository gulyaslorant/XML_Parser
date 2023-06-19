import openpyxl

workbook = openpyxl.load_workbook('Source/Test_Matrix.xlsx')

worksheet = workbook.active
counter=2
last_row = worksheet.max_row
while counter <= last_row:
    aktiv = worksheet.cell(row=counter, column=4).value
    set = worksheet.cell(row=counter, column=2).value
    anzahl = 0
    hauptzahl = 2
    while hauptzahl <= last_row:
        if set is worksheet.cell(row=hauptzahl, column=2).value:
            if aktiv is worksheet.cell(row=hauptzahl, column=4).value:
                anzahl +=1
        hauptzahl += 1
    worksheet.cell(row=counter, column=6).value = anzahl
    counter += 1
    print(anzahl)
workbook.save('chained.xlsx')
